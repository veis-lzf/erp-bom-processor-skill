import pandas as pd
import os
import glob
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BOM_COLUMNS = ['Item', 'Part NO.', 'Part Name', '规格型号', 'Quantity', 'Reference', 'PCB Footprint', 'Value']
DIFF_COLUMNS = ['Part NO.', 'Part Name', '规格型号', 'PCB Footprint', 'Value', '变化类型', '差异详情']
STATUS_UNMATCHED = '未匹配'


def extract_short_name(filepath):
    name = os.path.basename(filepath)
    name = re.sub(r'\.(BOM\.xlsx|BOM_processed\.xlsx|xlsx)$', '', name)
    name = name.replace('_processed', '')
    if len(name) > 30:
        name = name[:27] + '...'
    return name


def load_bom_files(file_paths):
    bom_data = {}
    for fp in file_paths:
        df = pd.read_excel(fp)
        name = extract_short_name(fp)
        df['_source'] = name
        df['_file'] = fp
        bom_data[name] = df
    return bom_data


def classify_part_no(part_no):
    if pd.isna(part_no) or str(part_no).strip() == '':
        return STATUS_UNMATCHED
    return str(part_no).strip()


def build_diff_data(bom_data):
    all_part_nos = set()
    bom_part_maps = {}

    for bom_name, df in bom_data.items():
        part_map = {}
        for _, row in df.iterrows():
            pn = classify_part_no(row['Part NO.'])
            if pn not in part_map:
                part_map[pn] = []
            part_map[pn].append(row)
        bom_part_maps[bom_name] = part_map
        all_part_nos.update(part_map.keys())

    return all_part_nos, bom_part_maps


def generate_diff_report(bom_data, bom_part_maps, all_part_nos):
    bom_names = list(bom_data.keys())
    summary_rows = []
    unique_rows = []
    common_diff_rows = []
    unmatched_rows = []

    for bom_name in bom_names:
        matched = 0
        unmatched = 0
        total = len(bom_data[bom_name])
        for pn, rows in bom_part_maps[bom_name].items():
            if pn == STATUS_UNMATCHED:
                unmatched += len(rows)
            else:
                matched += len(rows)
        summary_rows.append({
            'BOM名称': bom_name,
            '总物料数': total,
            '已匹配物料': matched,
            '未匹配物料': unmatched,
            '匹配率': f'{matched / total * 100:.1f}%' if total > 0 else '0%'
        })

    for pn in sorted(all_part_nos):
        if pn == STATUS_UNMATCHED:
            for bom_name in bom_names:
                if pn in bom_part_maps[bom_name]:
                    for row in bom_part_maps[bom_name][pn]:
                        unmatched_rows.append({
                            'BOM名称': bom_name,
                            'Part NO.': '',
                            'Part Name': row.get('Part Name', ''),
                            '规格型号': row.get('规格型号', ''),
                            'PCB Footprint': row.get('PCB Footprint', ''),
                            'Value': row.get('Value', ''),
                            'Quantity': row.get('Quantity', ''),
                            'Reference': row.get('Reference', '')
                        })
            continue

        appearing_boms = []
        for bom_name in bom_names:
            if pn in bom_part_maps[bom_name]:
                appearing_boms.append(bom_name)

        if len(appearing_boms) == 1:
            bom_name = appearing_boms[0]
            rows = bom_part_maps[bom_name][pn]
            for row in rows:
                unique_rows.append({
                    '所属BOM': bom_name,
                    'Part NO.': row.get('Part NO.', ''),
                    'Part Name': row.get('Part Name', ''),
                    '规格型号': row.get('规格型号', ''),
                    'PCB Footprint': row.get('PCB Footprint', ''),
                    'Value': row.get('Value', ''),
                    'Quantity': row.get('Quantity', ''),
                    'Reference': row.get('Reference', ''),
                    '变化类型': '独有物料'
                })
        else:
            first_bom = appearing_boms[0]
            first_rows = bom_part_maps[first_bom][pn]
            first_row = first_rows[0]

            has_diff = False
            diff_details = []

            for bom_name in appearing_boms[1:]:
                other_rows = bom_part_maps[bom_name][pn]
                other_row = other_rows[0]

                diffs_for_pair = []
                qty1 = first_row.get('Quantity', 0)
                qty2 = other_row.get('Quantity', 0)
                ref1 = str(first_row.get('Reference', ''))
                ref2 = str(other_row.get('Reference', ''))

                if qty1 != qty2:
                    diffs_for_pair.append(f"[{first_bom}]数量={qty1} vs [{bom_name}]数量={qty2}")

                if ref1 != ref2:
                    diffs_for_pair.append(f"[{first_bom}]位号={ref1} vs [{bom_name}]位号={ref2}")

                if diffs_for_pair:
                    has_diff = True
                    diff_details.append('; '.join(diffs_for_pair))

            if has_diff:
                common_diff_rows.append({
                    'Part NO.': first_row.get('Part NO.', ''),
                    'Part Name': first_row.get('Part Name', ''),
                    '规格型号': first_row.get('规格型号', ''),
                    'PCB Footprint': first_row.get('PCB Footprint', ''),
                    'Value': first_row.get('Value', ''),
                    '出现BOM': ', '.join(appearing_boms),
                    '变化类型': '数量/位号差异',
                    '差异详情': ' | '.join(diff_details)
                })

    return summary_rows, unique_rows, common_diff_rows, unmatched_rows


def apply_excel_style(ws, header_fill, header_font, border_style, row_fills):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_style

    for row_idx in range(2, ws.max_row + 1):
        fill = row_fills[row_idx % len(row_fills)]
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = border_style
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    for col_idx in range(1, ws.max_column + 1):
        max_width = 0
        for row_idx in range(1, ws.max_row + 1):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or '')
            char_width = 0
            for ch in cell_value:
                if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
                    char_width += 2
                else:
                    char_width += 1
            max_width = max(max_width, char_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 4, 45)


def write_diff_excel(output_path, summary_rows, unique_rows, common_diff_rows, unmatched_rows, bom_names):
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    border_style = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    row_fills = [
        PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
        PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
    ]

    ws_summary = wb.create_sheet('汇总统计')
    columns_summary = ['BOM名称', '总物料数', '已匹配物料', '未匹配物料', '匹配率']
    for col_idx, col_name in enumerate(columns_summary, 1):
        ws_summary.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row_data in enumerate(summary_rows, 2):
        for col_idx, col_name in enumerate(columns_summary, 1):
            ws_summary.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))
    apply_excel_style(ws_summary, header_fill, header_font, border_style, row_fills)

    if unique_rows:
        ws_unique = wb.create_sheet('独有物料')
        columns_unique = ['所属BOM', 'Part NO.', 'Part Name', '规格型号', 'PCB Footprint', 'Value', 'Quantity', 'Reference', '变化类型']
        for col_idx, col_name in enumerate(columns_unique, 1):
            ws_unique.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row_data in enumerate(unique_rows, 2):
            for col_idx, col_name in enumerate(columns_unique, 1):
                ws_unique.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))
        apply_excel_style(ws_unique, header_fill, header_font, border_style, row_fills)

        bom_color_map = {}
        bom_colors = ['FFD6E4', 'D6F5D6', 'D6E4FF', 'FFE4C4', 'E4D6FF']
        for i, name in enumerate(bom_names):
            bom_color_map[name] = PatternFill(start_color=bom_colors[i % len(bom_colors)], end_color=bom_colors[i % len(bom_colors)], fill_type='solid')

        for row_idx in range(2, ws_unique.max_row + 1):
            bom_cell = ws_unique.cell(row=row_idx, column=1)
            bom_name = str(bom_cell.value)
            if bom_name in bom_color_map:
                bom_cell.fill = bom_color_map[bom_name]

    if common_diff_rows:
        ws_common = wb.create_sheet('共有物料差异')
        columns_common = ['Part NO.', 'Part Name', '规格型号', 'PCB Footprint', 'Value', '出现BOM', '变化类型', '差异详情']
        for col_idx, col_name in enumerate(columns_common, 1):
            ws_common.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row_data in enumerate(common_diff_rows, 2):
            for col_idx, col_name in enumerate(columns_common, 1):
                ws_common.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))
        apply_excel_style(ws_common, header_fill, header_font, border_style, row_fills)

        diff_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        for row_idx in range(2, ws_common.max_row + 1):
            for col_idx in range(1, ws_common.max_column + 1):
                ws_common.cell(row=row_idx, column=col_idx).fill = diff_fill

    if unmatched_rows:
        ws_unmatched = wb.create_sheet('未匹配物料')
        columns_unmatched = ['BOM名称', 'Part NO.', 'Part Name', '规格型号', 'PCB Footprint', 'Value', 'Quantity', 'Reference']
        for col_idx, col_name in enumerate(columns_unmatched, 1):
            ws_unmatched.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row_data in enumerate(unmatched_rows, 2):
            for col_idx, col_name in enumerate(columns_unmatched, 1):
                ws_unmatched.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))
        apply_excel_style(ws_unmatched, header_fill, header_font, border_style, row_fills)

        warn_fill = PatternFill(start_color='FFD6D6', end_color='FFD6D6', fill_type='solid')
        for row_idx in range(2, ws_unmatched.max_row + 1):
            for col_idx in range(1, ws_unmatched.max_column + 1):
                ws_unmatched.cell(row=row_idx, column=col_idx).fill = warn_fill

    wb.save(output_path)
    return output_path


def bom_diff(file_paths, output_path):
    bom_data = load_bom_files(file_paths)
    all_part_nos, bom_part_maps = build_diff_data(bom_data)
    summary_rows, unique_rows, common_diff_rows, unmatched_rows = generate_diff_report(bom_data, bom_part_maps, all_part_nos)

    write_diff_excel(output_path, summary_rows, unique_rows, common_diff_rows, unmatched_rows, list(bom_data.keys()))

    print(f"\n=== BOM比对完成 ===")
    print(f"输出文件: {output_path}")
    print(f"\n--- 汇总统计 ---")
    for row in summary_rows:
        print(f"  {row['BOM名称']}: 总{row['总物料数']}项, 已匹配{row['已匹配物料']}项, 未匹配{row['未匹配物料']}项 ({row['匹配率']})")
    print(f"\n--- 差异详情 ---")
    print(f"  独有物料: {len(unique_rows)} 项")
    print(f"  共有物料差异: {len(common_diff_rows)} 项")
    print(f"  未匹配物料: {len(unmatched_rows)} 项")

    return output_path


if __name__ == '__main__':
    import sys

    OUTPUT_DIR = '04_output'
    DIFF_OUTPUT = os.path.join(OUTPUT_DIR, 'BOM差异清单.xlsx')

    if len(sys.argv) > 1:
        file_paths = sys.argv[1:]
    else:
        processed_files = glob.glob(os.path.join(OUTPUT_DIR, '*_processed.xlsx'))
        file_paths = sorted(processed_files)

    if not file_paths:
        print("未找到待比对的BOM文件")
        print("用法1: python bom_diff.py <file1> <file2> ...")
        print("用法2: python bom_diff.py (自动读取04_output下所有processed文件)")
        sys.exit(1)

    print(f"待比对文件: {len(file_paths)} 个")
    for fp in file_paths:
        print(f"  - {fp}")

    bom_diff(file_paths, DIFF_OUTPUT)